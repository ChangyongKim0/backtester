#-*-coding: utf-8
from request_wrapper import wrapped_requests as requests
from dart_crawler import DartCrawler
from utils import removeEmpty

from selenium.webdriver.common.alert import Alert
from datetime import datetime, timedelta
import xml.etree.ElementTree as ET
from selenium import webdriver
from bs4 import BeautifulSoup
from copy import deepcopy
import configparser
import subprocess
import openpyxl
import sqlite3
import pickle
import random
import time
import json
import glob
import sys
import re
import os

class FinancialCrawler:
    def __init__(self):
        self.save_dir = './data/financial'
        self.format_file = f'{self.save_dir}/format.xlsx'

        options = webdriver.ChromeOptions()
        options.add_argument('headless')
        self.driver = webdriver.Chrome('./chromedriver/chromedriver', options=options)
        self.driver.implicitly_wait(3)
    

    def getPageSource(self, post):
        jongmok_code = post['jongmok_code']
        url = f'https://navercomp.wisereport.co.kr/v2/company/c1010001.aspx?cmp_cd={jongmok_code}'
        print(f'[getPageSource] url : {url}')
        try:
            time.sleep(1)
            self.driver.get(url)
            self.driver.implicitly_wait(1)
            self.driver.find_element_by_id('cns_td21').click()
            page = self.driver.page_source
            return True, page
        except Exception as e:
            if "올바른 종목이 아닙니다" in e.args[0]:
                print(url, e.args[0])
                da = Alert(self.driver)
                da.accept()
            else:
                assert False
            return False, ""
    
    def saveFinancial(self, post):
        jongmok_code = post['jongmok_code']
        save_name = f'{self.save_dir}/{jongmok_code}.xlsx'

        result, page = self.getPageSource(post)
        if result == False:
            return
        
        soup = BeautifulSoup(page, 'html.parser')        
        table_list = soup.select('table.gHead01')
        table_item = None
        for table in table_list:
            caption = removeEmpty(table.select('caption.blind')[0].text)
            if "주요재무정보" == caption:
                table_item = table
                break
        if table_item is None:
            return

        wb = openpyxl.load_workbook(self.format_file)
        ws = wb['Sheet1']
        num_rows = ws.cell(row=1, column=2).value

        jongmok_name = soup.select('span.name')[0].text
        corp_WICS = removeEmpty(soup.select('td.td0101')[0].select('dt.line-left')[2].text.split(':')[1])
        ws.cell(row=2, column=2).value = jongmok_code
        ws.cell(row=3, column=2).value = jongmok_name
        ws.cell(row=4, column=2).value = corp_WICS

        tr_list = table_item.select('thead > tr')
        for tr_idx, tr_item in enumerate(tr_list):
            ws.cell(row=1, column=2).value += 1
            num_rows = ws.cell(row=1, column=2).value
            th_list = tr_item.select('th')
            if tr_idx == 0:
                ws.cell(row=4 + num_rows, column=1).value = removeEmpty(th_list[0].text)
                ws.cell(row=4 + num_rows, column=2).value = removeEmpty(th_list[0].text)
                ws.cell(row=4 + num_rows, column=6).value = removeEmpty(th_list[0].text)
            else:
                for th_idx in range(len(th_list)):
                    ws.cell(row=4 + num_rows, column=2 + th_idx).value = removeEmpty(th_list[th_idx].text)

        tr_list = table_item.select('tbody > tr')
        for tr_idx, tr_item in enumerate(tr_list):
            ws.cell(row=1, column=2).value += 1
            num_rows = ws.cell(row=1, column=2).value
            th_list = tr_item.select('th')
            ws.cell(row=4 + num_rows, column=1).value = removeEmpty(th_list[0].text)    
            td_list = tr_item.select('td')
            for td_idx in range(len(td_list)):
                ws.cell(row=4 + num_rows, column=2 + td_idx).value = removeEmpty(td_list[td_idx].text)
        wb.save(save_name)
        
    def saveFinancialList(self, post_list):
        for post_idx, post in enumerate(post_list):
            print(f"[saveFinancialList] {post_idx + 1} / {len(post_list)}")
            self.saveFinancial(post)

if __name__ == "__main__":
    financial_crawler = FinancialCrawler()
    dart_crawler = DartCrawler()

    post_list = dart_crawler.getUncheckedPostList()
    financial_crawler.saveFinancialList(post_list)
