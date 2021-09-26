import import_ipynb
import kw_algorithm_third as kw
import price_agent
import openpyxl
import readAndWriteExcel as excel
import os
import datetime as dt

# 변수 세팅

# 백테스팅 가정치
ouput_keys_assumptions = ['매수대기일수', '최대보유일수', '익절 퍼센트', '손절 퍼센트']
buy_waiting_period_list = [10] #공시 후 매수까지 avg_buy_price < target_date_low_price 되는지 기다리는 기간
holding_period_list = [70] #매수 후 매도까지 홀딩하는 기간. 이 기간동안 목표 수익률을 넘게 될 경우 기간 전이라도 매도
selling_plus_percent_list = [0.4] #홀딩기간 동안 익절 퍼센트
selling_minus_percent_list = [0.2] #홀딩기간 동안 손절 퍼센트
# sheet_splitter_indices = [0]
# file_splitter_indices = [1, 2]

# output 키 리스트 작성
# 20210619 00:00:00 최종 업데이트
# 공시날짜, 종목코드, 종목명, 매수날짜, 수정 종가 반영 매수가격, 총 매수금액, 매도날짜, 수정 종가 반영 매도가격, 총 매도 금액, 수익률, 비고(익절, 손절, 홀딩 기간 도래 매도)
output_keys = {'eng': ['gonsi_date', 'gonsi_url', 'jongmok_name', 'jongmok_code', 'jongmok_market' ,'buy_data', 'buy_price', 'buy_qty', 'sell_price', 'sell_qty', 'yield', 'sell_type'],
               'kor': ['공시날짜', '공시url', '종목명', '종목코드', '마켓종류', '매수날짜', '매수가격', '매수금액', '매도날짜', '매도가격', '매도금액', '수익률', '매도유형']}
lang = 'kor'
jongmok_code_index = 3 # 걍 종목코드 키 위치
finance_base_date_index = 0 # 재무정보를 가져올 기준날짜 키 위치
monitoring_base_date_index = 0 # 120일간의 추세를 확인할 기준날짜 키 위치

today = dt.datetime.now()
today = today.year,today.month, today.day

# output 경로 입력
filepath_out = '.'
filename_out = 'backtestor_{},{},{},{}_{}'.format(buy_waiting_period_list,holding_period_list,selling_plus_percent_list,selling_minus_percent_list, today)
sheetname_out_1 = 'Sheet1' # Sheet1 : 연도별 모든 재무데이터
sheetname_out_2 = 'Sheet2' # Sheet2 : 기준날짜 직전 재무데이터

# Finance 데이터 경로 입력
filepath_finance = 'financial'
filename_finance_sample = '000020' # 재무 키 리스트를 가져올 샘플 종목코드

# Price 데이터 경로 입력
filepath_price = '../kw/prices/all_daily_data_modified_close'
filepath_price_sample = '../kw/prices/all_daily_data'
filename_price_sample = '000020' # 영업일 리스트를 불러올 샘플 데이터

# 모니터링 기간 입력
monitoring_period = 120

# 기타 클래스/메소드 선언

# 걍 cartesian product 해주는거
def _cartProd2(a, b):
    prod = []
    for i in a:
        for j in b:
            try:
                k = [*i]
                k.append(j)
                prod.append(k)
            except:
                prod.append([i, j])
    return prod

def _cartProd(*argLists):
    prod = []
    for i, a in enumerate(argLists):
        if i > 0:
            prod = _cartProd2(prod, a)
        else:
            prod = a    
    return prod

set_assumptions = _cartProd(buy_waiting_period_list, holding_period_list, selling_plus_percent_list, selling_minus_percent_list)

# finance 업데이트용 클래스
class Finance:
    pivot_row = 6 # 엑셀 포맷이 바뀌면 수정 필요
    
    def __init__(self, filepath, sample, keys_assumptions, keys):
        self.filepath = filepath
        self.filename_sample = sample
        self.code_ws = {} # 중복으로 데이터를 불러오지 않기 위해 종목코드마다 불러온 워크시트를 할당. 
        # Sheet1 : 연도별 모든 재무데이터
        self.major_keys_1 = ['백테스팅 가정치'] + ['']*(len(keys_assumptions)-1) + ['기본정보'] + ['']*len(keys)
        self.minor_keys_1 = [*keys_assumptions]+[*keys]
        # Sheet2 : 기준날짜 직전 재무데이터
        self.major_keys_2 = [*self.major_keys_1]
        self.minor_keys_2 = [*self.minor_keys_1]
        self.error_list = []
        
        ws = excel.openfile(self.filepath, self.filename_sample)['Sheet1']
        self.code_ws[self.filename_sample] = ws
        self.year_cols = self._getYearCols(ws)
        self._appendKeys(ws)
        
    def updateAssumptions(self, bp, hp, sp, sm):
        self.assumptions = [bp, hp, sp, sm]
    
    def _appendKeys(self, ws):
        self.minor_keys_1.append(excel.cell(ws, 1, 4))
        self.minor_keys_2.append(excel.cell(ws, 1, 4))
        self.major_keys_2.append('직전년도 데이터')
        self.minor_keys_2.append('데이터 기준년도')
        
        for year in self.year_cols['list']:
            for row in range(self.pivot_row + 1, ws.max_row + 1):
                if row == self.pivot_row + 1:
                    self.major_keys_1.append('{}년 데이터'.format(str(year)))
                else:
                    self.major_keys_1.append('')
                self.minor_keys_1.append(excel.cell(ws, 1, row))
                
        for row in range(self.pivot_row + 1, ws.max_row + 1):
            self.major_keys_2.append('')
            self.minor_keys_2.append(excel.cell(ws, 1, row))
    
    # 모든 연도 목록과 해당 연도 열 위치 받아오기
    def _getYearCols(self, ws): 
        dict = {'list': []}
        for col in range(2, ws.max_column + 1):
            val = excel.cell(ws, col, self.pivot_row)
            if '/12' in val and not '/12(E)' in val:
                year = val.split('/')[0]
                dict['list'].append(int(year))
                dict[year] = col
        return dict

    # 주어진 종목코드에 해당하는 finance 데이터 추가
    def appendData(self, code, date, line):
        if not code in self.code_ws.keys():
            try:
                ws = excel.openfile(self.filepath, code)['Sheet1']
                self.code_ws[code] = ws
            except:
                self.error_list.append(code)
                return line, line
        else:
            ws = self.code_ws[code]
            
        line1 = self.assumptions + [*line]
        line2 = self.assumptions + [*line]
        line1.append(excel.cell(ws, 2, 4))
        line2.append(excel.cell(ws, 2, 4))
        
        for year in self.year_cols['list']:
            col = self.year_cols[str(year)]
            for row in range(self.pivot_row + 1, ws.max_row + 1):
                line1.append(excel.cell(ws, col, row))
        
        year_before = []
        year_now = int(date.split('-')[0])
        for year in self.year_cols['list']:
            if year < year_now:
                year_before.append(year)
        col = self.year_cols[str(max(year_before))]
        line2.append(max(year_before))
        for row in range(self.pivot_row + 1, ws.max_row + 1):
            line2.append(excel.cell(ws, col, row))
            
        return line1, line2
    
# 로그용 함수
def log(content):
    print('Log : {}'.format(content))

# 메인 코드

def run(for_test = False):
    wb, ws1, ws2 = excel.create(sheetname_out_1, sheetname_out_2)
    with_finance = Finance(filepath_finance, filename_finance_sample, ouput_keys_assumptions, output_keys[lang])
    keys_init = {'major1': with_finance.major_keys_1, 'major2': with_finance.major_keys_2, 'minor1': with_finance.minor_keys_1, 'minor2': with_finance.minor_keys_2}
    with_price = price_agent.Price(filepath_price, filepath_price_sample, filename_price_sample, monitoring_period, keys_init)
    excel.writeList(ws1, with_price.keys['major1'], is_first=True)
    excel.writeList(ws1, with_price.keys['minor1'])
    excel.writeList(ws2, with_price.keys['major2'], is_first=True)
    excel.writeList(ws2, with_price.keys['minor2'])
    
#     assert len(sheet_splitter_indices + file_splitter_indices) == len(set(sheet_splitter_indices + file_splitter_indices)), 'redundant index exists while splitting'

    for bp, hp, sp, sm in set_assumptions:
        with_finance.updateAssumptions(bp, hp, sp, sm)
        log('start backtesting with assumptions bp={}, hp={}, sp={}, sm={}'.format(bp, hp, sp, sm))
        if for_test: db_list = sample
        else: db_list = kw.run(bp, hp, sp, sm)
        print(db_list)
        log('the backtesting above ended')
        assert len(output_keys[lang]) == len(db_list[0]), 'different length between keys and values'
        for i, line in enumerate(db_list):
            line1_fi, line2_fi = with_finance.appendData(line[jongmok_code_index], line[finance_base_date_index], line)
            line1, line2 = with_price.appendData(line[jongmok_code_index], line[monitoring_base_date_index], with_price.nextDate(line[monitoring_base_date_index]), line1_fi, line2_fi)
            excel.writeList(ws1, line1)
            excel.writeList(ws2, line2)
        log('bactesting data successfully written')

    excel.save(filepath_out,filename_out, wb)
    log(with_finance.error_list)
    
if __name__ == '__main__':
    run(for_test = False)
    os.startfile('{}\{}.xlsx'.format(filepath_out.replace('/', 'W'), filename_out)) # 파일 열기 귀찮을때 쓰세욤